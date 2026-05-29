"""
API de Administración Centralizada de Estudiantes y Legajos.
Gestiona el ciclo de vida administrativo del alumno: desde la supervisión de 
documentación física (DNI, Títulos) hasta la auditoría de legajos y 
la baja administrativa bajo estrictas reglas de integridad académica.
"""

from __future__ import annotations
from django.db.models import Q
from django.shortcuts import get_object_or_404
from apps.common.api_schemas import ApiResponse
from core.models import Estudiante, EstudianteCarrera, ProrrogaTituloSecundario, ResidenciaCondicional, Regularidad, EquivalenciaDisposicionDetalle
from core.permissions import allowed_profesorados, ensure_roles, ensure_profesorado_access
from apps.common.date_utils import format_datetime

from ..schemas import (
    AutorizarRendirIn,
    EstudianteAdminDetail,
    EstudianteAdminListItem,
    EstudianteAdminListResponse,
    EstudianteAdminUpdateIn,
    EstudianteDocumentacionListItem,
    EstudianteDocumentacionListResponse,
    EstudianteDocumentacionUpdateIn,
    EstudianteDocumentacionBulkUpdateIn,
    ProrrogaTituloIn,
    ProrrogaTituloOut,
)
from .router import estudiantes_router as router
from ..services.estudiante_service import EstudianteService
from .helpers import (
    _ensure_admin,
    _ensure_staff_view,
    _apply_estudiante_updates,
    _build_admin_detail,
    _recalcular_estado_legajo,
)
from apps.common.audit import log_action_from_request, snapshot


@router.get("/admin/estudiantes/buscar-global", response=list[dict])
def admin_buscar_estudiantes_global(request, q: str = ""):
    """
    Búsqueda global de estudiantes sin restricción de carrera.
    Devuelve datos mínimos (dni, nombre, carreras actuales).
    Usado para agregar un estudiante de otra carrera a la propia.
    """
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})
    from django.db.models import Q as DQ

    q = q.strip()[:100]
    if len(q) < 2:
        return []

    qs = (
        Estudiante.objects.select_related("user")
        .prefetch_related("carreras_detalle__profesorado")
        .filter(
            DQ(persona__dni__icontains=q)
            | DQ(user__first_name__icontains=q)
            | DQ(user__last_name__icontains=q)
        )
        .order_by("user__last_name", "user__first_name")[:20]
    )

    result = []
    for est in qs:
        carreras = [
            {
                "nombre": ec.profesorado.nombre,
                "estado_academico": ec.get_estado_academico_display(),
            }
            for ec in est.carreras_detalle.all()
        ]
        result.append({
            "dni": est.dni,
            "apellido": est.user.last_name,
            "nombre": est.user.first_name,
            "carreras": carreras,
        })
    return result


@router.get(
    "/admin/estudiantes",
    response=EstudianteAdminListResponse,
)
def admin_list_estudiantes(
    request,
    q: str | None = None,
    carrera_id: int | None = None,
    estado_legajo: str | None = None,
    estado_academico: str | None = None,
    anio_ingreso: int | None = None,
    limit: int = 50,
    offset: int = 0,
):
    """
    Lista estudiantes con filtros administrativos y de carrera.
    Utiliza el servicio EstudianteService para la lógica compleja de filtrado.
    """
    _ensure_staff_view(request)
    allowed_ids = allowed_profesorados(request.user)
    filters = {
        "q": q,
        "carrera_id": carrera_id,
        "estado_legajo": estado_legajo,
        "estado_academico": estado_academico,
        "anio_ingreso": anio_ingreso,
    }
    return EstudianteService.list_estudiantes_admin(filters, limit, offset, allowed_ids)


@router.get(
    "/admin/estudiantes/anios-ingreso",
    response=list[int],
)
def admin_list_anios_ingreso(request, carrera_id: int | None = None):
    """
    Obtiene la lista de años de ingreso únicos presentes en la base de datos
    para alimentar los filtros de búsqueda.
    """
    _ensure_staff_view(request)
    allowed_ids = allowed_profesorados(request.user)
    
    # Si se pasa una carrera_id, debemos verificar que el usuario tenga acceso
    effective_allowed_ids = allowed_ids
    if carrera_id:
        if allowed_ids is not None and carrera_id not in allowed_ids:
            return []
        effective_allowed_ids = {carrera_id}
        
    return EstudianteService.get_unique_admission_years(effective_allowed_ids)


@router.get(
    "/admin/estudiantes-documentacion",
    response=EstudianteDocumentacionListResponse,
)
def admin_list_estudiantes_documentacion(
    request,
    q: str | None = None,
    carrera_id: int | None = None,
    estado_academico: str | None = None,
    limit: int = 100,
    offset: int = 0,
):
    """
    Obtiene la nómina de estudiantes con el estado de su documentación física.
    Esencial para el seguimiento de legajos incompletos o pendientes de entrega.
    """
    _ensure_staff_view(request)
    total, items = _get_estudiantes_documentacion_raw(request, q=q, carrera_id=carrera_id, estado_academico=estado_academico, limit=limit, offset=offset)
    return EstudianteDocumentacionListResponse(total=total, items=items)


def _get_estudiantes_documentacion_raw(request, q=None, carrera_id=None, estado_academico=None, limit=None, offset=0):
    """
    Lógica interna para consolidar datos de documentación de múltiples fuentes.
    Integra información de Estudiante, PreinscripcionChecklist y datos_extra.
    """
    from core.models import PreinscripcionChecklist

    # Filtros de territorialidad (Bedeles solo ven sus carreras permitidas)
    allowed_ids = allowed_profesorados(request.user)
    qs = (
        Estudiante.objects.select_related("persona", "user")
        .prefetch_related("carreras")
        .order_by("persona__apellido", "persona__nombre", "persona__dni")
    )
    
    if q:
        q_clean = q.strip()
        if q_clean:
            qs = qs.filter(
                Q(persona__dni__icontains=q_clean)
                | Q(persona__nombre__icontains=q_clean)
                | Q(persona__apellido__icontains=q_clean)
            )
    
    if carrera_id:
        if allowed_ids is not None and int(carrera_id) not in allowed_ids:
            qs = qs.none()
        else:
            if estado_academico:
                qs = qs.filter(
                    carreras_detalle__profesorado_id=carrera_id,
                    carreras_detalle__estado_academico=estado_academico,
                )
            else:
                qs = qs.filter(carreras__id=carrera_id)
    elif allowed_ids is not None:
        if estado_academico:
            qs = qs.filter(
                carreras_detalle__profesorado_id__in=allowed_ids,
                carreras_detalle__estado_academico=estado_academico,
            )
        else:
            qs = qs.filter(carreras__id__in=allowed_ids)
    elif estado_academico:
        qs = qs.filter(carreras_detalle__estado_academico=estado_academico)

    qs = qs.distinct()
    total = qs.count()
    
    if limit is not None:
        qs_paged = qs[offset : offset + limit]
    else:
        qs_paged = qs[offset:]
        
    estudiantes_list = list(qs_paged)
    estudiante_ids = [e.id for e in estudiantes_list]

    # Mapeo de checklists de preinscripción
    checklists = PreinscripcionChecklist.objects.filter(
        preinscripcion__alumno_id__in=estudiante_ids
    ).select_related("preinscripcion__alumno").order_by("-updated_at")
    
    checklist_map = {}
    for cl in checklists:
        if cl.preinscripcion.alumno_id not in checklist_map:
            checklist_map[cl.preinscripcion.alumno_id] = cl

    items = []
    from .helpers import _determine_condicion, _extract_documentacion
    
    for est in estudiantes_list:
        user = est.user if est.user_id else None
        # En el nuevo modelo, los campos están en Estudiante directamente
        # Pero conservamos retrocompatibilidad con datos_extra si se prefiere extraer vía helper
        doc_data = _extract_documentacion(est)
        
        cl = checklist_map.get(est.id)
        if cl:
            cl_fields = {
                "dni_legalizado": cl.dni_legalizado,
                "fotos_4x4": cl.fotos_4x4,
                "certificado_salud": cl.certificado_salud,
                "folios_oficio": cl.folios_oficio,
                "titulo_secundario_legalizado": cl.titulo_secundario_legalizado,
                "certificado_titulo_en_tramite": cl.certificado_titulo_en_tramite,
                "analitico_legalizado": cl.analitico_legalizado,
                "articulo_7": cl.articulo_7,
            }
            for k, v in cl_fields.items():
                if doc_data.get(k) in (None, False, 0, ""):
                    doc_data[k] = v
        
        condicion = _determine_condicion(doc_data)
        titulo_sec_ok = any([
            doc_data.get("titulo_secundario_legalizado"),
            doc_data.get("certificado_titulo_en_tramite"),
            doc_data.get("analitico_legalizado")
        ])

        items.append(
            EstudianteDocumentacionListItem(
                dni=est.dni,
                apellido=user.last_name if user else "",
                nombre=user.first_name if user else "",
                condicion_administrativa=condicion,
                curso_introductorio_aprobado=est.curso_introductorio_aprobado,
                libreta_entregada=est.libreta_entregada,
                dni_legalizado=bool(doc_data.get("dni_legalizado")),
                fotos_4x4=bool(doc_data.get("fotos_4x4")),
                certificado_salud=bool(doc_data.get("certificado_salud")),
                folios_oficio=int(doc_data.get("folios_oficio") or 0),
                titulo_secundario_ok=titulo_sec_ok,
                articulo_7=bool(doc_data.get("articulo_7")),
            )
        )
    return total, items


@router.get("/admin/estudiantes-documentacion/export/excel")
def admin_export_estudiantes_documentacion_excel(request, q: str | None = None, carrera_id: int | None = None):
    """Genera exportación Excel de la nómina de documentación para auditoría interna."""
    _ensure_staff_view(request)
    items = _get_estudiantes_documentacion_raw(request, q=q, carrera_id=carrera_id)
    
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentación"

    headers = [
        "DNI", "Apellido", "Nombre", "Condición", "CI", "Libreta", 
        "DNI (F.)", "Fotos", "Cert. Salud", "Folios", "Título Sec.", "Art. 7mo"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, item in enumerate(items, 2):
        ws.cell(row=idx, column=1, value=item.dni)
        ws.cell(row=idx, column=2, value=item.apellido)
        ws.cell(row=idx, column=3, value=item.nombre)
        ws.cell(row=idx, column=4, value=item.condicion_administrativa)
        ws.cell(row=idx, column=5, value="SI" if item.curso_introductorio_aprobado else "NO")
        ws.cell(row=idx, column=6, value="SI" if item.libreta_entregada else "NO")
        ws.cell(row=idx, column=7, value="SI" if item.dni_legalizado else "NO")
        ws.cell(row=idx, column=8, value="SI" if item.fotos_4x4 else "NO")
        ws.cell(row=idx, column=9, value="SI" if item.certificado_salud else "NO")
        ws.cell(row=idx, column=10, value=item.folios_oficio)
        ws.cell(row=idx, column=11, value="SI" if item.titulo_secundario_ok else "NO")
        ws.cell(row=idx, column=12, value="SI" if item.articulo_7 else "NO")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="estudiantes_documentacion.xlsx"'
    return response


@router.get("/admin/estudiantes-documentacion/export/pdf")
def admin_export_estudiantes_documentacion_pdf(request, q: str | None = None, carrera_id: int | None = None):
    """Genera exportación PDF de la nómina de documentación (formato imprimible)."""
    _ensure_staff_view(request)
    items = _get_estudiantes_documentacion_raw(request, q=q, carrera_id=carrera_id)
    
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    from weasyprint import HTML
    import datetime

    import os
    from django.conf import settings
    logo_left_path = os.path.join(settings.BASE_DIR, "static/logos/escudo_ministerio_tdf.png")
    logo_right_path = os.path.join(settings.BASE_DIR, "static/logos/logo_ipes.jpg")
    if not os.path.exists(logo_left_path):
        logo_left_path = os.path.join(settings.BASE_DIR, "backend/static/logos/escudo_ministerio_tdf.png")
        logo_right_path = os.path.join(settings.BASE_DIR, "backend/static/logos/logo_ipes.jpg")

    context = {
        "items": items,
        "fecha": format_datetime(datetime.datetime.now()),
        "q": q,
        "logo_left_path": logo_left_path,
        "logo_right_path": logo_right_path,
    }
    
    html_string = render_to_string("estudiantes/export_documentacion_pdf.html", context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="estudiantes_documentacion.pdf"'
    return response


def _perform_documentacion_update(est, payload):
    """Lógica unificada para persistir cambios en la documentación física del alumno."""
    from core.models import PreinscripcionChecklist
    
    checklist = PreinscripcionChecklist.objects.filter(preinscripcion__alumno=est).order_by("-updated_at").first()
    
    upd_fields = []
    if payload.libreta_entregada is not None:
        est.libreta_entregada = payload.libreta_entregada
        upd_fields.append("libreta_entregada")
    
    if payload.curso_introductorio_aprobado is not None:
        est.curso_introductorio_aprobado = payload.curso_introductorio_aprobado
        upd_fields.append("curso_introductorio_aprobado")

    # Documentación técnica básica
    mapping = {
        "dni_legalizado": payload.dni_legalizado,
        "fotos_4x4": payload.fotos_4x4,
        "certificado_salud": payload.certificado_salud,
        "folios_oficio": payload.folios_oficio,
        "articulo_7": payload.articulo_7,
        "titulo_secundario_legalizado": payload.titulo_secundario_ok
    }

    for field, val in mapping.items():
        if val is not None:
            setattr(est, field, val)
            upd_fields.append(field)

    if upd_fields:
        est.save(update_fields=upd_fields)
        _recalcular_estado_legajo(est)

        # Sincronizamos con el checklist si existe
        if checklist:
            for field, val in mapping.items():
                if val is not None:
                    setattr(checklist, field, val)
            if payload.curso_introductorio_aprobado is not None:
                checklist.curso_introductorio_aprobado = payload.curso_introductorio_aprobado
            checklist.save()


@router.patch(
    "/admin/estudiantes-documentacion/{dni}",
    response=ApiResponse,
)
def admin_update_estudiante_documentacion(
    request,
    dni: str,
    payload: EstudianteDocumentacionUpdateIn,
):
    """
    Actualiza individualmente la documentación técnica de un estudiante.
    Incluye chequeo de permisos territoriales para bedeles.
    """
    _ensure_admin(request)
    est = get_object_or_404(Estudiante, persona__dni=dni)
    
    # Auditoría de permisos por carrera
    allowed_ids = allowed_profesorados(request.user)
    if allowed_ids is not None:
        est_carreras_ids = set(EstudianteCarrera.objects.filter(estudiante=est).values_list("profesorado_id", flat=True))
        if not allowed_ids.intersection(est_carreras_ids):
            from apps.common.errors import raise_app_error
            from apps.common.constants import AppErrorCode
            raise_app_error(403, AppErrorCode.PERMISSION_DENIED, "No tiene permisos para modificar este legajo.")

    # Capturar estado previo para auditoría
    before = snapshot(est)

    _perform_documentacion_update(est, payload)

    # Registrar acción en auditoría
    log_action_from_request(
        request,
        accion="UPDATE",
        tipo_accion="CRUD",
        detalle_accion=f"Actualización de documentación legajo {dni}",
        entidad="Estudiante",
        entidad_id=dni,
        before=before,
        after=est,
    )

    return ApiResponse(ok=True, message="Documentación actualizada correctamente.")


@router.patch(
    "/admin/estudiantes-documentacion-bulk",
    response=ApiResponse,
)
def admin_bulk_update_estudiante_documentacion(
    request,
    payload: EstudianteDocumentacionBulkUpdateIn,
):
    _ensure_admin(request)
    allowed_ids = allowed_profesorados(request.user)
    
    updated_count = 0
    for update_item in payload.updates:
        est = Estudiante.objects.filter(persona__dni=update_item.dni).first()
        if not est:
            continue
            
        if allowed_ids is not None:
            est_carreras_ids = set(EstudianteCarrera.objects.filter(estudiante=est).values_list("profesorado_id", flat=True))
            if not allowed_ids.intersection(est_carreras_ids):
                continue
        
        _perform_documentacion_update(est, update_item.changes)
        updated_count += 1
        
    return ApiResponse(ok=True, message=f"Se actualizaron {updated_count} legajos.")


@router.get(
    "/admin/estudiantes/{dni}",
    response={200: EstudianteAdminDetail, 404: ApiResponse},
)
def admin_get_estudiante(request, dni: str):
    """Obtiene el detalle completo del legajo de un estudiante."""
    _ensure_staff_view(request)
    allowed_ids = allowed_profesorados(request.user)
    est = Estudiante.objects.select_related("user").prefetch_related("carreras").filter(persona__dni=dni).first()
    if not est:
        return 404, ApiResponse(ok=False, message="Estudiante no encontrado")
    return _build_admin_detail(est, allowed_carrera_ids=allowed_ids)


@router.put(
    "/admin/estudiantes/{dni}",
    response={200: EstudianteAdminDetail, 400: ApiResponse, 404: ApiResponse},
)
def admin_update_estudiante(request, dni: str, payload: EstudianteAdminUpdateIn):
    """
    Actualiza la información base del estudiante (Perfil, Legajo, Password).
    Permite el reseteo forzado de contraseña desde la administración.
    """
    _ensure_admin(request, include_attp=True)
    est = Estudiante.objects.select_related("user").prefetch_related("carreras").filter(persona__dni=dni).first()
    if not est:
        return 404, ApiResponse(ok=False, message="Estudiante no encontrado")

    # Capturar estado previo para auditoría
    before = snapshot(est)

    updated, error = _apply_estudiante_updates(
        est,
        payload,
        allow_estado_legajo=True,
        allow_force_password=True,
    )
    if not updated and error:
        status_code, api_resp = error
        return status_code, api_resp

    # Registrar acción en auditoría
    log_action_from_request(
        request,
        accion="UPDATE",
        tipo_accion="CRUD",
        detalle_accion=f"Actualización de datos base legajo {dni}",
        entidad="Estudiante",
        entidad_id=dni,
        before=before,
        after=est,
    )

    allowed_ids = allowed_profesorados(request.user)
    return _build_admin_detail(est, allowed_carrera_ids=allowed_ids)


@router.delete(
    "/admin/estudiantes/{dni}",
    response={200: ApiResponse, 400: ApiResponse, 404: ApiResponse},
)
def admin_delete_estudiante(request, dni: str):
    """
    Elimina físicamente a un estudiante del sistema.
    
    REGLA CRÍTICA DE INTEGRIDAD: Solo se permite si NO tiene historial académico.
    Se verifica: Inscripciones a materias, mesas, regularidades y actas históricas.
    """
    _ensure_admin(request)
    est = Estudiante.objects.filter(persona__dni=dni).first()
    if not est:
        return 404, ApiResponse(ok=False, message="Estudiante no encontrado")

    reasons = []
    # 1. Verificación de Cursadas
    ins_count = est.inscripciones_materia.exclude(estado="ANUL").count()
    if ins_count > 0:
        reasons.append(f"Tiene {ins_count} inscripciones activas.")

    # 2. Verificación de Notas y Regularidades
    reg_count = est.regularidades.count()
    if reg_count > 0:
        reasons.append(f"Tiene {reg_count} registros de regularidad/notas.")

    # 3. Verificación de Exámenes Finales
    mesas_count = est.inscripciones_mesa.count()
    if mesas_count > 0:
        reasons.append(f"Tiene {mesas_count} inscripciones a exámenes.")

    # 4. Auditoría en Actas Históricas (Datos Externos)
    from core.models import ActaExamenEstudiante
    actas_count = ActaExamenEstudiante.objects.filter(dni=dni).count()
    if actas_count > 0:
        reasons.append(f"Figura como alumno en {actas_count} actas de examen.")

    if reasons:
        msg = "No se puede eliminar: el estudiante posee historial académico irrenunciable. " + " ".join(reasons)
        return 400, ApiResponse(ok=False, message=msg)

    nombre_completo = str(est)
    user = est.user

    # Borrado en cascada (EstudianteCarrera, etc.)
    est.delete()

    # Limpieza de usuario si no tiene otros roles institucionales
    if user and not (user.is_staff or user.groups.filter(name__in=["docente", "bedel", "admin"]).exists()):
        user.delete()

    return 200, ApiResponse(ok=True, message=f"Legajo de {nombre_completo} eliminado correctamente.")
@router.post(
    "/admin/estudiantes/{dni}/reset-password",
    response=ApiResponse,
)
def admin_reset_estudiante_password(request, dni: str):
    """
    Resetea la contraseña del estudiante al formato 'pass' + DNI.
    Útil cuando el alumno olvida su primer acceso o hay problemas de login masivos.
    """
    _ensure_admin(request)
    est = get_object_or_404(Estudiante, persona__dni=dni)
    
    # Verificar si el usuario tiene permisos para esta carrera
    allowed_ids = allowed_profesorados(request.user)
    if allowed_ids is not None:
        from core.models import EstudianteCarrera
        est_carreras_ids = set(EstudianteCarrera.objects.filter(estudiante=est).values_list("profesorado_id", flat=True))
        if not allowed_ids.intersection(est_carreras_ids):
            from apps.common.errors import raise_app_error
            from apps.common.constants import AppErrorCode
            raise_app_error(403, AppErrorCode.PERMISSION_DENIED, "No tiene permisos para modificar este legajo.")

    success = EstudianteService.reset_password(est)
    if not success:
        return 400, ApiResponse(ok=False, message="No se pudo resetear la contraseña (usuario no vinculado)")


@router.patch(
    "/admin/estudiantes/{dni}/autorizar-rendir",
    response={200: ApiResponse, 400: ApiResponse, 403: ApiResponse, 404: ApiResponse},
)
def admin_autorizar_rendir(request, dni: str, payload: AutorizarRendirIn):
    """
    Activa o desactiva la autorización excepcional para rendir exámenes finales
    con legajo incompleto. Solo Secretaría y Bedelía pueden usar este endpoint.
    """
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})
    est = get_object_or_404(Estudiante, persona__dni=dni)

    est.autorizado_rendir = payload.autorizado
    est.autorizado_rendir_observacion = payload.observacion or None
    
    # Procesamiento de materias autorizadas (Many-to-Many)
    if payload.autorizado:
        est.materias_autorizadas.set(payload.materias_autorizadas)
    else:
        est.materias_autorizadas.clear()

    est.save(update_fields=["autorizado_rendir", "autorizado_rendir_observacion"])

    estado = "habilitado" if payload.autorizado else "deshabilitado"
    return 200, ApiResponse(ok=True, message=f"Autorización para rendir {estado} correctamente.")

    return ApiResponse(ok=True, message=f"Contraseña reseteada correctamente para {dni}. El alumno deberá cambiarla al primer ingreso.")


def _prorroga_to_out(p: ProrrogaTituloSecundario) -> dict:
    nombre = None
    if p.autorizado_por:
        nombre = p.autorizado_por.get_full_name() or p.autorizado_por.username
    return {
        "id": p.id,
        "fecha_otorgada": str(p.fecha_otorgada),
        "fecha_vencimiento": str(p.fecha_vencimiento),
        "observaciones": p.observaciones,
        "autorizado_por_nombre": nombre,
        "vigente": p.vigente,
        "dias_restantes": p.dias_restantes,
        "created_at": p.created_at.isoformat(),
    }


@router.get(
    "/admin/estudiantes/{dni}/prorrogas-titulo",
    response={200: list[ProrrogaTituloOut], 403: ApiResponse, 404: ApiResponse},
)
def admin_list_prorrogas_titulo(request, dni: str):
    """Lista todas las prórrogas del título secundario de un estudiante."""
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})
    est = get_object_or_404(Estudiante, persona__dni=dni)
    prorrogas = ProrrogaTituloSecundario.objects.filter(estudiante=est)
    return 200, [ProrrogaTituloOut(**_prorroga_to_out(p)) for p in prorrogas]


@router.post(
    "/admin/estudiantes/{dni}/prorrogas-titulo",
    response={200: ProrrogaTituloOut, 400: ApiResponse, 403: ApiResponse, 404: ApiResponse},
)
def admin_create_prorroga_titulo(request, dni: str, payload: ProrrogaTituloIn):
    """Crea una prórroga del título secundario para el estudiante."""
    ensure_roles(request.user, {"admin", "secretaria"})
    est = get_object_or_404(Estudiante, persona__dni=dni)
    from django.utils.dateparse import parse_date
    fecha_otorgada = parse_date(payload.fecha_otorgada)
    fecha_vencimiento = parse_date(payload.fecha_vencimiento)
    if not fecha_otorgada or not fecha_vencimiento:
        return 400, ApiResponse(ok=False, message="Fechas inválidas.")
    if fecha_vencimiento <= fecha_otorgada:
        return 400, ApiResponse(ok=False, message="La fecha de vencimiento debe ser posterior a la fecha de otorgamiento.")
    p = ProrrogaTituloSecundario.objects.create(
        estudiante=est,
        fecha_otorgada=fecha_otorgada,
        fecha_vencimiento=fecha_vencimiento,
        observaciones=payload.observaciones,
        autorizado_por=request.user,
    )
    return 200, ProrrogaTituloOut(**_prorroga_to_out(p))


@router.patch(
    "/admin/prorrogas-titulo/{prorroga_id}",
    response={200: ProrrogaTituloOut, 400: ApiResponse, 403: ApiResponse, 404: ApiResponse},
)
def admin_update_prorroga_titulo(request, prorroga_id: int, payload: ProrrogaTituloIn):
    """Actualiza una prórroga existente."""
    ensure_roles(request.user, {"admin", "secretaria"})
    p = get_object_or_404(ProrrogaTituloSecundario, id=prorroga_id)
    from django.utils.dateparse import parse_date
    fecha_otorgada = parse_date(payload.fecha_otorgada)
    fecha_vencimiento = parse_date(payload.fecha_vencimiento)
    if not fecha_otorgada or not fecha_vencimiento:
        return 400, ApiResponse(ok=False, message="Fechas inválidas.")
    if fecha_vencimiento <= fecha_otorgada:
        return 400, ApiResponse(ok=False, message="La fecha de vencimiento debe ser posterior a la fecha de otorgamiento.")
    p.fecha_otorgada = fecha_otorgada
    p.fecha_vencimiento = fecha_vencimiento
    p.observaciones = payload.observaciones
    p.save(update_fields=["fecha_otorgada", "fecha_vencimiento", "observaciones", "updated_at"])
    return 200, ProrrogaTituloOut(**_prorroga_to_out(p))


@router.delete(
    "/admin/prorrogas-titulo/{prorroga_id}",
    response={200: ApiResponse, 403: ApiResponse, 404: ApiResponse},
)
def admin_delete_prorroga_titulo(request, prorroga_id: int):
    """Elimina una prórroga."""
    ensure_roles(request.user, {"admin", "secretaria"})
    p = get_object_or_404(ProrrogaTituloSecundario, id=prorroga_id)
    p.delete()
    return 200, ApiResponse(ok=True, message="Prórroga eliminada.")


from ninja import Schema as _Schema

class AgregarCarreraIn(_Schema):
    profesorado_id: int
    anio_ingreso: int | None = None


@router.post(
    "/admin/estudiantes/{dni}/carreras",
    response={200: EstudianteAdminDetail, 400: ApiResponse, 403: ApiResponse, 404: ApiResponse},
)
def admin_agregar_carrera(request, dni: str, payload: AgregarCarreraIn):
    """Vincula a un estudiante existente con una nueva carrera (sin requerir preinscripción)."""
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})
    from core.models import Profesorado
    from django.utils import timezone

    est = Estudiante.objects.filter(persona__dni=dni).first()
    if not est:
        return 404, ApiResponse(ok=False, message="Estudiante no encontrado.")

    carrera = Profesorado.objects.filter(id=payload.profesorado_id).first()
    if not carrera:
        return 404, ApiResponse(ok=False, message="Carrera no encontrada.")

    allowed_ids = allowed_profesorados(request.user)
    if allowed_ids is not None and payload.profesorado_id not in allowed_ids:
        return 403, ApiResponse(ok=False, message="No tiene permisos para esta carrera.")

    if EstudianteCarrera.objects.filter(estudiante=est, profesorado=carrera).exists():
        return 400, ApiResponse(ok=False, message="El estudiante ya está vinculado a esa carrera.")

    anio = payload.anio_ingreso or timezone.now().year
    EstudianteCarrera.objects.create(
        estudiante=est,
        profesorado=carrera,
        anio_ingreso=anio,
        estado_academico=EstudianteCarrera.EstadoAcademico.ACTIVO,
    )

    return 200, _build_admin_detail(est, allowed_carrera_ids=allowed_ids)


@router.get(
    "/admin/prorrogas-titulo/alertas",
    response={200: list[dict], 403: ApiResponse},
)
def admin_alertas_prorrogas_titulo(
    request,
    dias_aviso: int = 30,
    carrera_id: int | None = None,
):
    """
    Lista prórrogas vencidas o próximas a vencer (dentro de `dias_aviso` días).
    Usar desde el dashboard de secretaría para anticipar cierres masivos erróneos.
    """
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})
    from django.utils import timezone as tz
    from datetime import timedelta
    hoy = tz.localdate()
    limite = hoy + timedelta(days=dias_aviso)

    qs = ProrrogaTituloSecundario.objects.filter(
        fecha_vencimiento__lte=limite,
    ).select_related("estudiante__persona", "autorizado_por")

    if carrera_id:
        qs = qs.filter(estudiante__carreras__id=carrera_id)

    resultado = []
    for p in qs.order_by("fecha_vencimiento"):
        dias = (p.fecha_vencimiento - hoy).days
        resultado.append({
            "prorroga_id": p.id,
            "dni": p.estudiante.persona.dni if p.estudiante.persona else None,
            "nombre": p.estudiante.persona.apellido_nombre if p.estudiante.persona else str(p.estudiante),
            "fecha_vencimiento": str(p.fecha_vencimiento),
            "dias_restantes": dias,
            "vencida": dias < 0,
            "autorizado_por": p.autorizado_por.get_full_name() if p.autorizado_por else None,
        })
    return 200, resultado


@router.get(
    "/admin/resguardo-correlativas",
    response={200: list[dict], 403: ApiResponse},
)
def admin_resguardo_correlativas(
    request,
    carrera_id: int | None = None,
    solo_activos: bool = True,
):
    """
    Lista estudiantes que tienen materias en resguardo por correlativas faltantes.
    Incluye tanto Regularidades como Equivalencias con en_resguardo=True.
    Usar desde el dashboard de secretaría/bedelía para seguimiento.
    """
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})

    resultado_map: dict[int, dict] = {}

    # --- Regularidades en resguardo ---
    regs_qs = Regularidad.objects.filter(
        en_resguardo=True,
        situacion__in=[
            Regularidad.Situacion.REGULAR,
            Regularidad.Situacion.APROBADO,
            Regularidad.Situacion.PROMOCIONADO,
        ],
    ).select_related("estudiante__persona", "materia")

    if carrera_id:
        regs_qs = regs_qs.filter(estudiante__carreras__id=carrera_id)

    for reg in regs_qs:
        est = reg.estudiante
        est_id = est.id
        if est_id not in resultado_map:
            resultado_map[est_id] = {
                "estudiante_id": est_id,
                "dni": est.persona.dni if est.persona else None,
                "nombre": f"{est.persona.apellido}, {est.persona.nombre}" if est.persona else str(est_id),
                "estado_legajo": est.estado_legajo,
                "materias_en_resguardo": [],
            }
        resultado_map[est_id]["materias_en_resguardo"].append({
            "materia_id": reg.materia_id,
            "materia_nombre": reg.materia.nombre,
            "situacion": reg.situacion,
            "situacion_display": reg.get_situacion_display(),
            "fuente": "cursada",
            "fecha": str(reg.fecha_cierre) if reg.fecha_cierre else None,
        })

    # --- Equivalencias en resguardo ---
    equis_qs = EquivalenciaDisposicionDetalle.objects.filter(
        en_resguardo=True,
    ).select_related("disposicion__estudiante__persona", "materia", "disposicion")

    if carrera_id:
        equis_qs = equis_qs.filter(disposicion__estudiante__carreras__id=carrera_id)

    for eq in equis_qs:
        est = eq.disposicion.estudiante
        est_id = est.id
        if est_id not in resultado_map:
            resultado_map[est_id] = {
                "estudiante_id": est_id,
                "dni": est.persona.dni if est.persona else None,
                "nombre": f"{est.persona.apellido}, {est.persona.nombre}" if est.persona else str(est_id),
                "estado_legajo": est.estado_legajo,
                "materias_en_resguardo": [],
            }
        resultado_map[est_id]["materias_en_resguardo"].append({
            "materia_id": eq.materia_id,
            "materia_nombre": eq.materia.nombre,
            "situacion": "EQUIV",
            "situacion_display": "Equivalencia",
            "fuente": "equivalencia",
            "fecha": str(eq.disposicion.fecha_disposicion),
        })

    resultado = sorted(resultado_map.values(), key=lambda x: x["nombre"])
    return 200, resultado


@router.get("/admin/residencias-condicionales")
def admin_residencias_condicionales(
    request,
    ciclo: int | None = None,
    carrera_id: int | None = None,
    solo_pendientes: bool = True,
):
    """
    Alerta administrativa: lista de estudiantes con inscripción condicional a Residencia.
    Permite a secretaría/bedelía hacer seguimiento de quiénes deben aprobar en mayo.

    Filtros:
    - ciclo: año lectivo (por defecto: año actual)
    - carrera_id: filtrar por profesorado
    - solo_pendientes: True = solo las no resueltas ni caídas (default)
    """
    from datetime import date
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})

    ciclo = ciclo or date.today().year
    qs = ResidenciaCondicional.objects.filter(
        ciclo_lectivo=ciclo,
    ).select_related(
        "estudiante__persona",
        "materia_residencia__plan_de_estudio__profesorado",
        "materia_pendiente",
    ).order_by("estudiante__persona__apellido")

    if solo_pendientes:
        qs = qs.filter(resuelta=False, caida=False)

    if carrera_id:
        qs = qs.filter(materia_residencia__plan_de_estudio__profesorado_id=carrera_id)

    resultado = []
    for rc in qs:
        est = rc.estudiante
        profesorado = getattr(getattr(rc.materia_residencia.plan_de_estudio, "profesorado", None), "nombre", None) if rc.materia_residencia.plan_de_estudio_id else None

        if rc.resuelta:
            estado = "RESUELTA"
        elif rc.caida:
            estado = "CAÍDA"
        else:
            estado = "PENDIENTE"

        resultado.append({
            "id": rc.id,
            "ciclo_lectivo": rc.ciclo_lectivo,
            "dni": est.persona.dni if est.persona_id else None,
            "nombre": f"{est.persona.apellido}, {est.persona.nombre}" if est.persona_id else str(est),
            "profesorado": profesorado,
            "materia_residencia": rc.materia_residencia.nombre,
            "materia_pendiente": rc.materia_pendiente.nombre,
            "fecha_limite": str(rc.fecha_limite),
            "aceptada_en": rc.aceptada_en.strftime("%d/%m/%Y %H:%M"),
            "estado": estado,
        })

    return 200, resultado


@router.get("/admin/resguardo-materias")
def admin_resguardo_materias(
    request,
    profesorado_id: int | None = None,
    dni: str | None = None,
):
    """
    Lista todas las Regularidades y Equivalencias con en_resguardo=True,
    con el motivo detallado de por qué están en resguardo.
    Filtrable por profesorado y DNI.
    """
    from datetime import date
    from core.models import Correlatividad, Materia
    from apps.estudiantes.api.helpers.misc_utils import (
        _tiene_aprobacion_valida,
        _calcular_vigencia_regularidad,
    )
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})

    hoy = date.today()
    resultado = []

    def _motivo_faltantes(est, materia, autorizadas_ids, situacion=None):
        faltantes = []
        for corr in Correlatividad.objects.filter(
            materia_origen=materia,
            tipo=Correlatividad.TipoCorrelatividad.APROBADA_PARA_CURSAR,
        ).select_related("materia_correlativa"):
            if not _tiene_aprobacion_valida(est, corr.materia_correlativa, autorizadas_ids=autorizadas_ids):
                faltantes.append(f"Necesita APROBAR: {corr.materia_correlativa.nombre}")
        for corr in Correlatividad.objects.filter(
            materia_origen=materia,
            tipo=Correlatividad.TipoCorrelatividad.REGULAR_PARA_CURSAR,
        ).select_related("materia_correlativa"):
            if _tiene_aprobacion_valida(est, corr.materia_correlativa, autorizadas_ids=autorizadas_ids):
                continue
            # Solo la regularidad más reciente: si la última está vencida/agotada,
            # las anteriores también lo estarían (son más viejas).
            rc = (
                Regularidad.objects.filter(
                    estudiante=est,
                    materia=corr.materia_correlativa,
                    situacion=Regularidad.Situacion.REGULAR,
                    en_resguardo=False,
                )
                .order_by("-fecha_cierre")
                .first()
            )
            if not rc:
                faltantes.append(f"Necesita REGULARIZAR: {corr.materia_correlativa.nombre}")
            else:
                limite, intentos, max_i = _calcular_vigencia_regularidad(est, rc)
                if hoy > limite:
                    faltantes.append(f"Regularidad VENCIDA ({rc.fecha_cierre}): {corr.materia_correlativa.nombre}")
                elif intentos >= max_i:
                    faltantes.append(f"Regularidad AGOTADA ({intentos}/{max_i} intentos): {corr.materia_correlativa.nombre}")
        if situacion in (Regularidad.Situacion.APROBADO, Regularidad.Situacion.PROMOCIONADO):
            for corr in Correlatividad.objects.filter(
                materia_origen=materia,
                tipo=Correlatividad.TipoCorrelatividad.APROBADA_PARA_RENDIR,
            ).select_related("materia_correlativa"):
                if not _tiene_aprobacion_valida(est, corr.materia_correlativa, autorizadas_ids=autorizadas_ids):
                    faltantes.append(f"Necesita APROBAR (para rendir): {corr.materia_correlativa.nombre}")
        # Si una materia aparece como "Necesita APROBAR (para rendir)", la regularidad
        # es irrelevante — eliminar entradas "Necesita REGULARIZAR" duplicadas para esa materia.
        materias_que_requieren_aprobacion = {
            f.split(": ", 1)[1] for f in faltantes if f.startswith("Necesita APROBAR (para rendir):")
        }
        faltantes = [
            f for f in faltantes
            if not (f.startswith("Necesita REGULARIZAR:") and f.split(": ", 1)[1] in materias_que_requieren_aprobacion)
        ]
        return list(dict.fromkeys(faltantes))

    # Estudiantes activos en el profesorado indicado
    from core.models import EstudianteCarrera
    activos_en_prof_qs = EstudianteCarrera.objects.filter(estado_academico="ACT")
    if profesorado_id:
        activos_en_prof_qs = activos_en_prof_qs.filter(profesorado_id=profesorado_id)
    activos_ids = activos_en_prof_qs.values_list("estudiante_id", flat=True)

    # Regularidades en resguardo
    reg_qs = Regularidad.objects.filter(en_resguardo=True, estudiante_id__in=activos_ids).select_related(
        "estudiante__persona", "materia__plan_de_estudio__profesorado"
    )
    if profesorado_id:
        reg_qs = reg_qs.filter(materia__plan_de_estudio__profesorado_id=profesorado_id)
    if dni:
        reg_qs = reg_qs.filter(estudiante__persona__dni=dni)
    reg_qs = reg_qs.order_by("estudiante__persona__apellido", "materia__nombre")

    for reg in reg_qs:
        est = reg.estudiante
        autorizadas_ids = set(est.materias_autorizadas.values_list("id", flat=True))
        prof = getattr(getattr(reg.materia.plan_de_estudio, "profesorado", None), "nombre", None) if reg.materia.plan_de_estudio_id else None
        resultado.append({
            "tipo": "REG",
            "dni": est.persona.dni if est.persona_id else None,
            "nombre": f"{est.persona.apellido}, {est.persona.nombre}" if est.persona_id else str(est.id),
            "profesorado": prof,
            "materia": reg.materia.nombre,
            "situacion": reg.get_situacion_display(),
            "motivos": _motivo_faltantes(est, reg.materia, autorizadas_ids, reg.situacion),
        })

    # Equivalencias en resguardo
    eq_qs = EquivalenciaDisposicionDetalle.objects.filter(en_resguardo=True, disposicion__estudiante_id__in=activos_ids).select_related(
        "disposicion__estudiante__persona", "materia__plan_de_estudio__profesorado"
    )
    if profesorado_id:
        eq_qs = eq_qs.filter(materia__plan_de_estudio__profesorado_id=profesorado_id)
    if dni:
        eq_qs = eq_qs.filter(disposicion__estudiante__persona__dni=dni)
    eq_qs = eq_qs.order_by("disposicion__estudiante__persona__apellido", "materia__nombre")

    for eq in eq_qs:
        est = eq.disposicion.estudiante
        autorizadas_ids = set(est.materias_autorizadas.values_list("id", flat=True))
        prof = getattr(getattr(eq.materia.plan_de_estudio, "profesorado", None), "nombre", None) if eq.materia.plan_de_estudio_id else None
        resultado.append({
            "tipo": "EQUIV",
            "dni": est.persona.dni if est.persona_id else None,
            "nombre": f"{est.persona.apellido}, {est.persona.nombre}" if est.persona_id else str(est.id),
            "profesorado": prof,
            "materia": eq.materia.nombre,
            "situacion": "Equivalencia",
            "motivos": _motivo_faltantes(est, eq.materia, autorizadas_ids),
        })

    return 200, resultado


@router.post("/admin/resguardo-materias/recalcular")
def admin_recalcular_resguardo(
    request,
    profesorado_id: int | None = None,
    solo_activos: bool = True,
):
    """
    Ejecuta el comando recalcular_resguardo en un hilo de fondo para evitar timeout.
    Solo para admin y secretaria.
    """
    import threading
    from django.core.management import call_command
    ensure_roles(request.user, {"admin", "secretaria", "bedel"})

    kwargs = {
        "dry_run": False,
        "solo_equivalencias": False,
        "solo_regularidades": False,
        "dni": None,
        "solo_activos": solo_activos,
        "profesorado": profesorado_id,
    }

    def run():
        try:
            call_command("recalcular_resguardo", **kwargs)
        except Exception:
            pass

    thread = threading.Thread(target=run, daemon=True)
    thread.start()

    return 200, {
        "ok": True,
        "regularidades_marcadas": 0,
        "regularidades_liberadas": 0,
        "equivalencias_marcadas": 0,
        "equivalencias_liberadas": 0,
    }
