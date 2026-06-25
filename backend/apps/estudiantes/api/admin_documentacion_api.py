"""
API de Administración Centralizada de Estudiantes y Legajos.
Gestiona el ciclo de vida administrativo del alumno: desde la supervisión de
documentación física (DNI, Títulos) hasta la auditoría de legajos y
la baja administrativa bajo estrictas reglas de integridad académica.
"""

from __future__ import annotations

from django.db.models import Q
from django.shortcuts import get_object_or_404

from apps.common.api_schemas import ApiResponse
from apps.common.audit import log_action_from_request, snapshot
from apps.common.date_utils import format_datetime
from core.models import (
    EquivalenciaDisposicionDetalle,
    Estudiante,
    EstudianteCarrera,
    ProrrogaTituloSecundario,
    Regularidad,
    ResidenciaCondicional,
)
from core.permissions import allowed_profesorados, require

from ..schemas import (
    AutorizarRendirIn,
    EstudianteAdminDetail,
    EstudianteAdminListItem,
    EstudianteAdminListResponse,
    EstudianteAdminUpdateIn,
    EstudianteDocumentacionBulkUpdateIn,
    EstudianteDocumentacionListItem,
    EstudianteDocumentacionListResponse,
    EstudianteDocumentacionUpdateIn,
    ProrrogaTituloIn,
    ProrrogaTituloOut,
)
from ..services.estudiante_service import EstudianteService
from .helpers import (
    _apply_estudiante_updates,
    _build_admin_detail,
    _recalcular_estado_legajo,
)
from .router import estudiantes_router as router


@router.get(
    "/admin/estudiantes-documentacion",
    response=EstudianteDocumentacionListResponse,
)
def admin_list_estudiantes_documentacion(
    request,
    q: str | None = None,
    carrera_id: int | None = None,
    estado_academico: str | None = None,
    limit: int = 100,
    offset: int = 0,
):
    """
    Obtiene la nómina de estudiantes con el estado de su documentación física.
    Esencial para el seguimiento de legajos incompletos o pendientes de entrega.
    """
    require(request.user, "ver_estudiantes")
    total, items = _get_estudiantes_documentacion_raw(
        request, q=q, carrera_id=carrera_id, estado_academico=estado_academico, limit=limit, offset=offset
    )
    return EstudianteDocumentacionListResponse(total=total, items=items)


def _get_estudiantes_documentacion_raw(request, q=None, carrera_id=None, estado_academico=None, limit=None, offset=0):
    """
    Lógica interna para consolidar datos de documentación de múltiples fuentes.
    Integra información de Estudiante, PreinscripcionChecklist y datos_extra.
    """
    from core.models import PreinscripcionChecklist

    # Defensivo: el frontend puede mandar "undefined" o "null" como string
    if q and q.strip().lower() in ("undefined", "null"):
        q = None

    # Filtros de territorialidad (Bedeles solo ven sus carreras permitidas)
    allowed_ids = allowed_profesorados(request.user)
    qs = (
        Estudiante.objects.select_related("persona", "user")
        .prefetch_related("carreras")
        .order_by("persona__apellido", "persona__nombre", "persona__dni")
    )

    if q:
        q_clean = q.strip()
        if q_clean:
            qs = qs.filter(
                Q(persona__dni__icontains=q_clean)
                | Q(persona__nombre__icontains=q_clean)
                | Q(persona__apellido__icontains=q_clean)
            )

    if carrera_id:
        if allowed_ids is not None and int(carrera_id) not in allowed_ids:
            qs = qs.none()
        else:
            if estado_academico:
                qs = qs.filter(
                    carreras_detalle__profesorado_id=carrera_id,
                    carreras_detalle__estado_academico=estado_academico,
                )
            else:
                qs = qs.filter(carreras__id=carrera_id)
    elif allowed_ids is not None:
        if estado_academico:
            qs = qs.filter(
                carreras_detalle__profesorado_id__in=allowed_ids,
                carreras_detalle__estado_academico=estado_academico,
            )
        else:
            qs = qs.filter(carreras__id__in=allowed_ids)
    elif estado_academico:
        qs = qs.filter(carreras_detalle__estado_academico=estado_academico)

    qs = qs.distinct()
    total = qs.count()

    if limit is not None:
        qs_paged = qs[offset : offset + limit]
    else:
        qs_paged = qs[offset:]

    estudiantes_list = list(qs_paged)
    estudiante_ids = [e.id for e in estudiantes_list]

    # Mapeo de checklists de preinscripción
    checklists = (
        PreinscripcionChecklist.objects.filter(preinscripcion__alumno_id__in=estudiante_ids)
        .select_related("preinscripcion__alumno")
        .order_by("-updated_at")
    )

    checklist_map = {}
    for cl in checklists:
        if cl.preinscripcion.alumno_id not in checklist_map:
            checklist_map[cl.preinscripcion.alumno_id] = cl

    # Pre-cachear EstudianteCarrera por estudiante para evitar N+1
    from core.models import EstudianteCarrera as _EC

    ec_qs = _EC.objects.filter(estudiante_id__in=estudiante_ids)
    ec_map: dict[int, list] = {}
    for ec in ec_qs:
        ec_map.setdefault(ec.estudiante_id, []).append(ec)

    EC_DOC_FIELDS = [
        "dni_legalizado",
        "fotos_4x4",
        "certificado_salud",
        "folios_oficio",
        "titulo_secundario_legalizado",
        "certificado_titulo_en_tramite",
        "analitico_legalizado",
        "articulo_7",
        "certificado_alumno_regular_sec",
        "adeuda_materias",
        "es_certificacion_docente",
    ]

    items = []
    from .helpers import _determine_condicion, _extract_documentacion

    for est in estudiantes_list:
        user = est.user if est.user_id else None
        persona = est.persona
        doc_data = _extract_documentacion(est)

        # Mergear con datos de EstudianteCarrera (fuente primaria de documentación)
        for ec in ec_map.get(est.id, []):
            for field in EC_DOC_FIELDS:
                if not doc_data.get(field):
                    doc_data[field] = getattr(ec, field, False)

        cl = checklist_map.get(est.id)
        if cl:
            cl_fields = {
                "dni_legalizado": cl.dni_legalizado,
                "fotos_4x4": cl.fotos_4x4,
                "certificado_salud": cl.certificado_salud,
                "folios_oficio": cl.folios_oficio,
                "titulo_secundario_legalizado": cl.titulo_secundario_legalizado,
                "certificado_titulo_en_tramite": cl.certificado_titulo_en_tramite,
                "analitico_legalizado": cl.analitico_legalizado,
                "articulo_7": cl.articulo_7,
            }
            for k, v in cl_fields.items():
                if doc_data.get(k) in (None, False, 0, ""):
                    doc_data[k] = v

        condicion = _determine_condicion(doc_data)
        titulo_sec_ok = any(
            [
                doc_data.get("titulo_secundario_legalizado"),
                doc_data.get("certificado_titulo_en_tramite"),
                doc_data.get("analitico_legalizado"),
            ]
        )

        items.append(
            EstudianteDocumentacionListItem(
                dni=est.dni,
                apellido=persona.apellido if persona else "",
                nombre=persona.nombre if persona else "",
                email=user.email if user else "",
                condicion_administrativa=condicion,
                curso_introductorio_aprobado=est.curso_introductorio_aprobado,
                libreta_entregada=est.libreta_entregada,
                dni_legalizado=bool(doc_data.get("dni_legalizado")),
                fotos_4x4=bool(doc_data.get("fotos_4x4")),
                certificado_salud=bool(doc_data.get("certificado_salud")),
                folios_oficio=int(doc_data.get("folios_oficio") or 0),
                titulo_secundario_ok=titulo_sec_ok,
                articulo_7=bool(doc_data.get("articulo_7")),
            )
        )
    return total, items


@router.get("/admin/estudiantes-documentacion/export/excel")
def admin_export_estudiantes_documentacion_excel(
    request, q: str | None = None, carrera_id: int | None = None, estado_academico: str | None = None
):
    """Genera exportación Excel de la nómina de documentación para auditoría interna."""
    require(request.user, "ver_estudiantes")
    _total, items = _get_estudiantes_documentacion_raw(
        request, q=q, carrera_id=carrera_id, estado_academico=estado_academico
    )

    from io import BytesIO

    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentación"

    headers = [
        "DNI",
        "Apellido",
        "Nombre",
        "Correo",
        "Condición",
        "CI",
        "Libreta",
        "DNI (F.)",
        "Fotos",
        "Cert. Salud",
        "Folios",
        "Título Sec.",
        "Art. 7mo",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, item in enumerate(items, 2):
        ws.cell(row=idx, column=1, value=item.dni)
        ws.cell(row=idx, column=2, value=item.apellido)
        ws.cell(row=idx, column=3, value=item.nombre)
        ws.cell(row=idx, column=4, value=item.email)
        ws.cell(row=idx, column=5, value=item.condicion_administrativa)
        ws.cell(row=idx, column=6, value="SI" if item.curso_introductorio_aprobado else "NO")
        ws.cell(row=idx, column=7, value="SI" if item.libreta_entregada else "NO")
        ws.cell(row=idx, column=8, value="SI" if item.dni_legalizado else "NO")
        ws.cell(row=idx, column=9, value="SI" if item.fotos_4x4 else "NO")
        ws.cell(row=idx, column=10, value="SI" if item.certificado_salud else "NO")
        ws.cell(row=idx, column=11, value=item.folios_oficio)
        ws.cell(row=idx, column=12, value="SI" if item.titulo_secundario_ok else "NO")
        ws.cell(row=idx, column=13, value="SI" if item.articulo_7 else "NO")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="estudiantes_documentacion.xlsx"'
    return response


@router.get("/admin/estudiantes-documentacion/export/pdf")
def admin_export_estudiantes_documentacion_pdf(
    request, q: str | None = None, carrera_id: int | None = None, estado_academico: str | None = None
):
    """Genera exportación PDF de la nómina de documentación (formato imprimible)."""
    require(request.user, "ver_estudiantes")
    _total, items = _get_estudiantes_documentacion_raw(
        request, q=q, carrera_id=carrera_id, estado_academico=estado_academico
    )

    import datetime
    import os

    from django.conf import settings
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from weasyprint import HTML

    logo_left_path = os.path.join(settings.BASE_DIR, "static/logos/escudo_ministerio_tdf.png")
    logo_right_path = os.path.join(settings.BASE_DIR, "static/logos/logo_ipes.jpg")
    if not os.path.exists(logo_left_path):
        logo_left_path = os.path.join(settings.BASE_DIR, "backend/static/logos/escudo_ministerio_tdf.png")
        logo_right_path = os.path.join(settings.BASE_DIR, "backend/static/logos/logo_ipes.jpg")

    context = {
        "items": items,
        "fecha": format_datetime(datetime.datetime.now()),
        "q": q,
        "logo_left_path": logo_left_path,
        "logo_right_path": logo_right_path,
    }

    html_string = render_to_string("estudiantes/export_documentacion_pdf.html", context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="estudiantes_documentacion.pdf"'
    return response


def _perform_documentacion_update(est, payload):
    """Lógica unificada para persistir cambios en la documentación física del alumno."""
    from core.models import PreinscripcionChecklist

    checklist = PreinscripcionChecklist.objects.filter(preinscripcion__alumno=est).order_by("-updated_at").first()

    upd_fields = []
    if payload.libreta_entregada is not None:
        est.libreta_entregada = payload.libreta_entregada
        upd_fields.append("libreta_entregada")

    if payload.curso_introductorio_aprobado is not None:
        est.curso_introductorio_aprobado = payload.curso_introductorio_aprobado
        upd_fields.append("curso_introductorio_aprobado")

    # Documentación técnica básica
    mapping = {
        "dni_legalizado": payload.dni_legalizado,
        "fotos_4x4": payload.fotos_4x4,
        "certificado_salud": payload.certificado_salud,
        "folios_oficio": payload.folios_oficio,
        "articulo_7": payload.articulo_7,
        "titulo_secundario_legalizado": payload.titulo_secundario_ok,
    }

    for field, val in mapping.items():
        if val is not None:
            setattr(est, field, val)
            upd_fields.append(field)

    if upd_fields:
        est.save(update_fields=upd_fields)

        # Sincronizar los mismos campos en todas las EstudianteCarrera del alumno
        # (_recalcular_estado_legajo_ec lee de EC, no de Estudiante)
        ec_doc_fields = [f for f in mapping if mapping[f] is not None]
        if ec_doc_fields:
            for ec in est.carreras_detalle.all():
                ec_changed = False
                for field in ec_doc_fields:
                    val = mapping[field]
                    if getattr(ec, field, None) != val:
                        setattr(ec, field, val)
                        ec_changed = True
                if (
                    payload.curso_introductorio_aprobado is not None
                    and ec.curso_introductorio_aprobado != payload.curso_introductorio_aprobado
                ):
                    ec.curso_introductorio_aprobado = payload.curso_introductorio_aprobado
                    ec_changed = True
                if ec_changed:
                    ec.save(
                        update_fields=ec_doc_fields
                        + (["curso_introductorio_aprobado"] if payload.curso_introductorio_aprobado is not None else [])
                    )

        # Sincronizamos con el checklist si existe
        if checklist:
            for field, val in mapping.items():
                if val is not None:
                    setattr(checklist, field, val)
            if payload.curso_introductorio_aprobado is not None:
                checklist.curso_introductorio_aprobado = payload.curso_introductorio_aprobado
            checklist.save()

        # Recalcular DESPUÉS de que tanto EC como checklist estén actualizados
        _recalcular_estado_legajo(est)


@router.patch(
    "/admin/estudiantes-documentacion/{dni}",
    response=ApiResponse,
)
def admin_update_estudiante_documentacion(
    request,
    dni: str,
    payload: EstudianteDocumentacionUpdateIn,
):
    """
    Actualiza individualmente la documentación técnica de un estudiante.
    Incluye chequeo de permisos territoriales para bedeles.
    """
    require(request.user, "editar_documentacion")
    est = get_object_or_404(Estudiante, persona__dni=dni)

    # Auditoría de permisos por carrera
    allowed_ids = allowed_profesorados(request.user)
    if allowed_ids is not None:
        est_carreras_ids = set(
            EstudianteCarrera.objects.filter(estudiante=est).values_list("profesorado_id", flat=True)
        )
        if not allowed_ids.intersection(est_carreras_ids):
            from apps.common.constants import AppErrorCode
            from apps.common.errors import raise_app_error

            raise_app_error(403, AppErrorCode.PERMISSION_DENIED, "No tiene permisos para modificar este legajo.")

    # Capturar estado previo para auditoría
    before = snapshot(est)

    _perform_documentacion_update(est, payload)

    # Registrar acción en auditoría
    log_action_from_request(
        request,
        accion="UPDATE",
        tipo_accion="CRUD",
        detalle_accion=f"Actualización de documentación legajo {dni}",
        entidad="Estudiante",
        entidad_id=dni,
        before=before,
        after=est,
    )

    return ApiResponse(ok=True, message="Documentación actualizada correctamente.")


@router.patch(
    "/admin/estudiantes-documentacion-bulk",
    response=ApiResponse,
)
def admin_bulk_update_estudiante_documentacion(
    request,
    payload: EstudianteDocumentacionBulkUpdateIn,
):
    require(request.user, "editar_documentacion")
    allowed_ids = allowed_profesorados(request.user)

    updated_count = 0
    for update_item in payload.updates:
        est = Estudiante.objects.filter(persona__dni=update_item.dni).first()
        if not est:
            continue

        if allowed_ids is not None:
            est_carreras_ids = set(
                EstudianteCarrera.objects.filter(estudiante=est).values_list("profesorado_id", flat=True)
            )
            if not allowed_ids.intersection(est_carreras_ids):
                continue

        _perform_documentacion_update(est, update_item.changes)
        updated_count += 1

    return ApiResponse(ok=True, message=f"Se actualizaron {updated_count} legajos.")
