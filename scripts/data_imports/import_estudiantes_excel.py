#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import difflib
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any


MOJIBAKE_MARKERS = ("Ã", "Â", "�")


def has_mojibake(value: str | None) -> bool:
    if not value:
        return False
    return any(marker in value for marker in MOJIBAKE_MARKERS)


def maybe_fix_mojibake(value: str | None) -> str | None:
    if value is None:
        return None
    v = str(value)
    if not has_mojibake(v):
        return v
    try:
        fixed = v.encode("latin1", errors="ignore").decode("utf-8", errors="ignore")
    except Exception:
        return v
    return fixed if fixed else v


def normalize_label(value: str | None) -> str:
    if not value:
        return ""
    value = maybe_fix_mojibake(str(value).strip()) or ""
    value = unicodedata.normalize("NFKD", value)
    value = value.encode("ascii", "ignore").decode("ascii")
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    text = str(value).strip()
    text = maybe_fix_mojibake(text) or text
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_date(value: Any) -> dt.date | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime(value: Any) -> dt.datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min)
    text = clean_text(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def clean_dni(value: Any) -> str:
    digits = re.sub(r"\D+", "", clean_text(value))
    return digits[:10]


def clean_phone(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    return digits[:20]


def clean_email(value: Any) -> str:
    email = clean_text(value).lower()
    if not email or "@" not in email:
        return ""
    return email[:254]


@dataclass
class ExcelRow:
    dni: str
    cuil: str
    fecha_nacimiento: dt.date | None
    telefono: str
    domicilio: str
    email: str
    ciudadnac: str
    provnac: str
    paisnac: str
    titsec: str
    estsec: str
    titsuperior: str
    estsup: str
    carrera_raw: str
    anio_ingreso: int | None
    cohorte: str
    source_row: int
    source_dt: dt.datetime | None


ALIASES = {
    "dni": ["dni", "documento", "nro documento", "numero documento"],
    "cuil": ["cuil", "cuit", "cuil/cuit"],
    "fecha_nacimiento": ["fecha nacimiento", "fecha_nacimiento", "f nac", "nacimiento"],
    "telefono": ["telefono", "tel", "celular", "telefono celular"],
    "domicilio": ["domicilio", "direccion"],
    "email": ["email", "correo", "mail"],
    "ciudadnac": ["ciudadnac", "ciudad nacimiento", "lugar nacimiento", "localidad nacimiento"],
    "provnac": ["provnac", "provincia nacimiento"],
    "paisnac": ["paisnac", "pais nacimiento"],
    "titsec": ["titsec", "titulo secundario", "secundario", "secundario completo"],
    "estsec": ["estsec", "establecimiento secundario", "colegio secundario"],
    "titsuperior": ["titsuperior", "titulo superior", "terciario superior"],
    "estsup": ["estsup", "establecimiento superior", "institucion superior"],
    "carrera": ["carrera", "profesorado", "carrera inscripcion", "inscripcion"],
    "fechahora": ["fechahora", "fecha hora", "timestamp", "created_at", "fecha"],
    "anio_ingreso": ["anio ingreso", "año ingreso", "anio", "año"],
    "cohorte": ["cohorte"],
}


def load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        print(f"ERROR: falta dependencia openpyxl: {exc}", file=sys.stderr)
        print("Instala en el entorno: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean_text(h) for h in next(rows)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        record: dict[str, Any] = {"__rownum__": idx}
        for i, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[i] if i < len(row) else None
        out.append(record)
    return out


def detect_columns(records: list[dict[str, Any]]) -> dict[str, str]:
    if not records:
        return {}
    headers = list(records[0].keys())
    norm_to_real = {normalize_label(h): h for h in headers}
    selected: dict[str, str] = {}
    for logical, aliases in ALIASES.items():
        for alias in aliases:
            k = norm_to_real.get(normalize_label(alias))
            if k:
                selected[logical] = k
                break
    return selected


def build_excel_rows(records: list[dict[str, Any]], cols: dict[str, str]) -> list[ExcelRow]:
    out: list[ExcelRow] = []
    for rec in records:
        dni = clean_dni(rec.get(cols.get("dni", ""), ""))
        if not dni:
            continue
        anio_raw = clean_text(rec.get(cols.get("anio_ingreso", ""), ""))
        anio_ingreso = int(anio_raw) if anio_raw.isdigit() and 1900 <= int(anio_raw) <= 2100 else None
        out.append(
            ExcelRow(
                dni=dni,
                cuil=clean_text(rec.get(cols.get("cuil", ""), ""))[:32],
                fecha_nacimiento=parse_date(rec.get(cols.get("fecha_nacimiento", ""), "")),
                telefono=clean_phone(rec.get(cols.get("telefono", ""), "")),
                domicilio=clean_text(rec.get(cols.get("domicilio", ""), ""))[:255],
                email=clean_email(rec.get(cols.get("email", ""), "")),
                ciudadnac=clean_text(rec.get(cols.get("ciudadnac", ""), ""))[:255],
                provnac=clean_text(rec.get(cols.get("provnac", ""), ""))[:255],
                paisnac=clean_text(rec.get(cols.get("paisnac", ""), ""))[:255],
                titsec=clean_text(rec.get(cols.get("titsec", ""), ""))[:255],
                estsec=clean_text(rec.get(cols.get("estsec", ""), ""))[:255],
                titsuperior=clean_text(rec.get(cols.get("titsuperior", ""), ""))[:255],
                estsup=clean_text(rec.get(cols.get("estsup", ""), ""))[:255],
                carrera_raw=clean_text(rec.get(cols.get("carrera", ""), ""))[:255],
                anio_ingreso=anio_ingreso,
                cohorte=clean_text(rec.get(cols.get("cohorte", ""), ""))[:32],
                source_row=int(rec.get("__rownum__", 0)),
                source_dt=parse_datetime(rec.get(cols.get("fechahora", ""), "")),
            )
        )
    return out


def dedupe_latest(rows: list[ExcelRow]) -> dict[str, ExcelRow]:
    best: dict[str, ExcelRow] = {}
    for row in rows:
        current = best.get(row.dni)
        if current is None:
            best[row.dni] = row
            continue
        a = row.source_dt or dt.datetime.min
        b = current.source_dt or dt.datetime.min
        if (a, row.source_row) >= (b, current.source_row):
            best[row.dni] = row
    return best


def connect_db(args: argparse.Namespace):
    try:
        import pymysql
    except Exception as exc:  # pragma: no cover
        print(f"ERROR: falta dependencia pymysql: {exc}", file=sys.stderr)
        print("Instala en el entorno: pip install pymysql", file=sys.stderr)
        sys.exit(2)
    return pymysql.connect(
        host=args.db_host,
        port=args.db_port,
        user=args.db_user,
        password=args.db_password,
        database=args.db_name,
        charset="utf8mb4",
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )


def fetch_students(cur) -> dict[str, dict[str, Any]]:
    cur.execute(
        """
        SELECT
          e.id AS estudiante_id,
          e.user_id,
          e.dni,
          e.fecha_nacimiento,
          e.telefono,
          e.domicilio,
          e.datos_extra,
          u.email,
          u.last_name,
          u.first_name
        FROM core_estudiante e
        JOIN auth_user u ON u.id = e.user_id
        """
    )
    out: dict[str, dict[str, Any]] = {}
    for row in cur.fetchall():
        dni = clean_dni(row["dni"])
        if not dni:
            continue
        datos_extra = row.get("datos_extra") or {}
        if isinstance(datos_extra, str):
            try:
                datos_extra = json.loads(datos_extra)
            except json.JSONDecodeError:
                datos_extra = {}
        row["datos_extra"] = datos_extra
        out[dni] = row
    return out


def fetch_profesorados(cur) -> list[dict[str, Any]]:
    cur.execute("SELECT id, nombre, activo FROM core_profesorado")
    return list(cur.fetchall())


def resolve_profesorado(
    carrera_raw: str,
    professorados: list[dict[str, Any]],
    manual_map: dict[str, str],
    allow_fuzzy: bool,
) -> tuple[int | None, str]:
    if not carrera_raw:
        return None, "empty"
    normalized = normalize_label(carrera_raw)
    if not normalized:
        return None, "empty"

    if normalized in manual_map:
        target = manual_map[normalized]
        if str(target).isdigit():
            return int(target), "manual-id"
        target_norm = normalize_label(target)
        for p in professorados:
            if normalize_label(p["nombre"]) == target_norm:
                return int(p["id"]), "manual-name"
        return None, "manual-miss"

    exact: dict[str, int] = {normalize_label(p["nombre"]): int(p["id"]) for p in professorados}
    if normalized in exact:
        return exact[normalized], "exact"

    if not allow_fuzzy:
        return None, "unresolved"

    names = list(exact.keys())
    guess = difflib.get_close_matches(normalized, names, n=1, cutoff=0.93)
    if not guess:
        return None, "unresolved"
    return exact[guess[0]], "fuzzy"


def load_manual_map(path: Path | None) -> dict[str, str]:
    if not path:
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, str] = {}
    for k, v in data.items():
        out[normalize_label(str(k))] = str(v)
    return out


def maybe_snapshot(cur, prefix: str) -> None:
    if not prefix:
        return
    base = re.sub(r"[^a-zA-Z0-9_]+", "_", prefix)[:40]
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    t_est = f"{base}_core_estudiante_{stamp}"
    t_car = f"{base}_core_estudiantecarrera_{stamp}"
    t_usr = f"{base}_auth_user_{stamp}"
    cur.execute(f"CREATE TABLE `{t_est}` AS SELECT * FROM core_estudiante")
    cur.execute(f"CREATE TABLE `{t_car}` AS SELECT * FROM core_estudiantecarrera")
    cur.execute(f"CREATE TABLE `{t_usr}` AS SELECT id, email FROM auth_user")


def apply_mojibake_fix_existing(cur) -> None:
    cur.execute(
        """
        UPDATE core_estudiante
        SET domicilio = CONVERT(CAST(CONVERT(domicilio USING latin1) AS BINARY) USING utf8mb4)
        WHERE domicilio REGEXP 'Ã|Â|�'
        """
    )
    cur.execute(
        """
        UPDATE core_estudiante
        SET datos_extra = JSON_SET(
            datos_extra,
            '$.ciudadnac', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.ciudadnac')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.ciudadnac')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.ciudadnac'))),
            '$.provnac', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.provnac')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.provnac')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.provnac'))),
            '$.paisnac', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.paisnac')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.paisnac')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.paisnac'))),
            '$.titsec', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsec')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsec')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsec'))),
            '$.estsec', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsec')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsec')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsec'))),
            '$.titsuperior', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsuperior')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsuperior')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsuperior'))),
            '$.estsup', IF(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsup')) REGEXP 'Ã|Â|�', CONVERT(CAST(CONVERT(JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsup')) USING latin1) AS BINARY) USING utf8mb4), JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsup')))
        )
        WHERE CONCAT_WS('|',
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.ciudadnac')),
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.provnac')),
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.paisnac')),
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsec')),
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsec')),
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.titsuperior')),
          JSON_UNQUOTE(JSON_EXTRACT(datos_extra,'$.estsup'))
        ) REGEXP 'Ã|Â|�'
        """
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Importador seguro de datos de estudiantes desde Excel.")
    parser.add_argument("--excel", required=True, help="Ruta al XLSX de entrada.")
    parser.add_argument("--out-dir", default="Temporales", help="Directorio de salida para reportes.")
    parser.add_argument("--apply", action="store_true", help="Aplica cambios en DB. Por defecto simula.")
    parser.add_argument("--confirm", default="", help='Confirmacion obligatoria para --apply: escribir "APLICAR".')
    parser.add_argument("--career-map", default="", help="JSON opcional para mapear carrera_excel -> profesorado.")
    parser.add_argument("--allow-fuzzy-career", action="store_true", help="Permite fuzzy matching para carreras.")
    parser.add_argument("--fix-existing-mojibake", action="store_true", help="Corrige mojibake preexistente en DB.")
    parser.add_argument("--snapshot-prefix", default="", help="Si se define, crea tablas snapshot antes de aplicar.")
    parser.add_argument("--db-host", default=os.getenv("DB_HOST", "127.0.0.1"))
    parser.add_argument("--db-port", type=int, default=int(os.getenv("DB_PORT", "3306")))
    parser.add_argument("--db-user", default=os.getenv("DB_USER", "ipes_user"))
    parser.add_argument("--db-password", default=os.getenv("DB_PASSWORD", ""))
    parser.add_argument("--db-name", default=os.getenv("DB_NAME", "ipes6"))
    args = parser.parse_args()

    if args.apply and args.confirm != "APLICAR":
        print('ERROR: para aplicar debes pasar --confirm APLICAR', file=sys.stderr)
        return 2

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: no existe excel: {excel_path}", file=sys.stderr)
        return 2

    run_ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = Path(args.out_dir) / f"import_estudiantes_{run_ts}"
    run_dir.mkdir(parents=True, exist_ok=True)

    records = load_xlsx_rows(excel_path)
    cols = detect_columns(records)
    rows = build_excel_rows(records, cols)
    by_dni = dedupe_latest(rows)
    manual_map = load_manual_map(Path(args.career_map)) if args.career_map else {}

    con = connect_db(args)
    unresolved: list[tuple[str, str]] = []
    stats: dict[str, int] = {
        "excel_rows": len(rows),
        "excel_unique_dni": len(by_dni),
        "db_dni": 0,
        "interseccion": 0,
        "solo_excel": 0,
        "solo_db": 0,
        "updates_estudiante": 0,
        "updates_email": 0,
        "updates_datos_extra": 0,
        "career_links": 0,
        "career_unresolved": 0,
    }

    changes_preview: list[list[str]] = []
    only_excel: list[str] = []
    only_db: list[str] = []

    try:
        with con.cursor() as cur:
            db_students = fetch_students(cur)
            profs = fetch_profesorados(cur)
            db_dnis = set(db_students.keys())
            ex_dnis = set(by_dni.keys())

            stats["db_dni"] = len(db_dnis)
            stats["interseccion"] = len(ex_dnis & db_dnis)
            stats["solo_excel"] = len(ex_dnis - db_dnis)
            stats["solo_db"] = len(db_dnis - ex_dnis)
            only_excel = sorted(ex_dnis - db_dnis)
            only_db = sorted(db_dnis - ex_dnis)

            for dni in sorted(ex_dnis & db_dnis):
                src = by_dni[dni]
                dst = db_students[dni]

                est_updates: dict[str, Any] = {}
                if src.fecha_nacimiento and src.fecha_nacimiento != dst.get("fecha_nacimiento"):
                    est_updates["fecha_nacimiento"] = src.fecha_nacimiento
                if src.telefono and src.telefono != (dst.get("telefono") or ""):
                    est_updates["telefono"] = src.telefono
                if src.domicilio and src.domicilio != (dst.get("domicilio") or ""):
                    est_updates["domicilio"] = src.domicilio

                extra = dict(dst.get("datos_extra") or {})
                extra_updates = {}
                for k in ("ciudadnac", "provnac", "paisnac", "titsec", "estsec", "titsuperior", "estsup"):
                    v = getattr(src, k)
                    if v and v != clean_text(extra.get(k, "")):
                        extra_updates[k] = v
                for k, v in extra_updates.items():
                    extra[k] = v

                email_update = src.email and src.email != (dst.get("email") or "")

                prof_id, reason = resolve_profesorado(src.carrera_raw, profs, manual_map, args.allow_fuzzy_career)
                if src.carrera_raw and prof_id is None:
                    stats["career_unresolved"] += 1
                    unresolved.append((dni, src.carrera_raw))

                if est_updates:
                    stats["updates_estudiante"] += 1
                if extra_updates:
                    stats["updates_datos_extra"] += 1
                if email_update:
                    stats["updates_email"] += 1
                if prof_id:
                    stats["career_links"] += 1

                if len(changes_preview) < 200 and (est_updates or extra_updates or email_update or prof_id):
                    changes_preview.append(
                        [
                            dni,
                            dst.get("last_name", ""),
                            dst.get("first_name", ""),
                            json.dumps(est_updates, ensure_ascii=False),
                            json.dumps(extra_updates, ensure_ascii=False),
                            src.email if email_update else "",
                            src.carrera_raw,
                            str(prof_id or ""),
                            reason,
                        ]
                    )

                if args.apply:
                    if est_updates:
                        set_expr = ", ".join(f"{k}=%s" for k in est_updates)
                        params = list(est_updates.values()) + [dst["estudiante_id"]]
                        cur.execute(f"UPDATE core_estudiante SET {set_expr} WHERE id=%s", params)
                    if extra_updates:
                        cur.execute("UPDATE core_estudiante SET datos_extra=%s WHERE id=%s", (json.dumps(extra, ensure_ascii=False), dst["estudiante_id"]))
                    if email_update:
                        cur.execute("UPDATE auth_user SET email=%s WHERE id=%s", (src.email, dst["user_id"]))
                    if prof_id:
                        cur.execute(
                            """
                            INSERT INTO core_estudiantecarrera
                              (estudiante_id, profesorado_id, anio_ingreso, cohorte, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, NOW(), NOW())
                            ON DUPLICATE KEY UPDATE
                              anio_ingreso = COALESCE(VALUES(anio_ingreso), anio_ingreso),
                              cohorte = CASE WHEN VALUES(cohorte) IS NULL OR VALUES(cohorte) = '' THEN cohorte ELSE VALUES(cohorte) END,
                              updated_at = NOW()
                            """,
                            (dst["estudiante_id"], prof_id, src.anio_ingreso, src.cohorte),
                        )

            if args.apply:
                if args.snapshot_prefix:
                    maybe_snapshot(cur, args.snapshot_prefix)
                if args.fix_existing_mojibake:
                    apply_mojibake_fix_existing(cur)
                con.commit()
            else:
                con.rollback()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

    (run_dir / "stats.json").write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    (run_dir / "dni_only_excel.txt").write_text("\n".join(only_excel) + ("\n" if only_excel else ""), encoding="utf-8")
    (run_dir / "dni_only_db.txt").write_text("\n".join(only_db) + ("\n" if only_db else ""), encoding="utf-8")
    if unresolved:
        (run_dir / "career_unresolved.csv").write_text(
            "dni,carrera_excel\n" + "\n".join(f"{dni},{json.dumps(carr, ensure_ascii=False)}" for dni, carr in unresolved) + "\n",
            encoding="utf-8",
        )

    with (run_dir / "preview_changes.csv").open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            [
                "dni",
                "apellido",
                "nombre",
                "estudiante_updates",
                "datos_extra_updates",
                "email_nuevo",
                "carrera_excel",
                "profesorado_id",
                "carrera_match_reason",
            ]
        )
        writer.writerows(changes_preview)

    summary = [
        f"Modo: {'APPLY' if args.apply else 'SIMULACION'}",
        f"Run: {run_dir}",
        f"Excel filas con DNI: {stats['excel_rows']}",
        f"Excel DNI unicos: {stats['excel_unique_dni']}",
        f"DB DNI: {stats['db_dni']}",
        f"Interseccion: {stats['interseccion']}",
        f"DNI solo Excel: {stats['solo_excel']}",
        f"DNI solo DB: {stats['solo_db']}",
        f"Filas con update core_estudiante: {stats['updates_estudiante']}",
        f"Filas con update datos_extra: {stats['updates_datos_extra']}",
        f"Filas con update email: {stats['updates_email']}",
        f"Vinculos carrera (resueltos): {stats['career_links']}",
        f"Carreras no resueltas: {stats['career_unresolved']}",
    ]
    if args.apply:
        summary.append("Aplicado con transaccion: COMMIT")
    else:
        summary.append("No se aplicaron cambios (ROLLBACK).")
    (run_dir / "summary.txt").write_text("\n".join(summary) + "\n", encoding="utf-8")
    print("\n".join(summary))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
